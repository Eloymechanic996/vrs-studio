import json
from copy import deepcopy
from datetime import datetime
from pathlib import Path

from kivy.app import App
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner
from kivy.uix.tabbedpanel import TabbedPanel, TabbedPanelItem
from kivy.uix.textinput import TextInput
from openpyxl import Workbook


TRACK_STATES = [
    "Green Flag",
    "Yellow Flag",
    "Red Flag",
    "Safety Car",
    "Virtual Safety Car",
    "Pit Lane",
]

EVENT_TYPES = [
    "Incidencia",
    "Pit In",
    "Pit Out",
    "Safety Car Sale",
    "Safety Car Entra",
    "Bandera",
    "Cambio de clima",
]

SESSION_TYPES = [
    "Race",
    "Test",
    "Free Practice",
    "Qualifying",
]

MODALITIES = [
    "Karting",
    "Monoplazas",
    "Rally",
    "Turismos",
    "Endurance",
    "Off-Road",
]

WEATHER_CONDITIONS = [
    "Dry",
    "Cloudy",
    "Light Rain",
    "Rain",
    "Heavy Rain",
    "Windy",
    "Cold",
    "Hot",
]


def today_text():
    return datetime.now().strftime("%Y-%m-%d")


DEFAULT_FORM = {
    "modalidad": "Karting",
    "categoria": "",
    "tipo_sesion": "Test",
    "campeonato": "",
    "equipo": "",
    "evento": "",
    "circuito": "",
    "fecha": today_text(),
    "piloto": "",
    "dorsal": "",
    "mecanico": "Jose",
    "chasis": "",
    "motor_nombre": "",
    "clima": "Dry",
    "clima_pista": "Dry",
    "estado_pista": TRACK_STATES[0],
    "tipo_evento": EVENT_TYPES[0],
    "session_notes": "",
    "timing_notes": "",
    "distancia_delante": "",
    "distancia_detras": "",
    "caida_izq": "",
    "caida_der": "",
    "caster_izq": "",
    "caster_der": "",
    "reactividad": "",
    "avance": "",
    "neumatico": "",
    "llanta": "",
    "buje": "",
    "ancho_trasero": "",
    "altura_eje": "",
    "pinon": "",
    "corona": "",
    "relacion": "-",
    "tipo_eje": "",
    "presion_del_izq": "",
    "presion_del_der": "",
    "presion_tras_izq": "",
    "presion_tras_der": "",
    "bujia": "",
    "carburacion": "",
    "desarrollo": "",
    "temperatura_motor": "",
    "aguja_altas": "",
    "aguja_bajas": "",
}


DEFAULT_TIMING = {
    "timer_running": False,
    "start_time": None,
    "elapsed_before_start": 0.0,
    "last_lap_elapsed": 0.0,
    "lap_number": 1,
    "pending_lap_status": None,
    "rows": [],
}


class MobileRoot(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", spacing=dp(10), padding=dp(10), **kwargs)
        self.form_values = deepcopy(DEFAULT_FORM)
        self.timing_state = deepcopy(DEFAULT_TIMING)
        self.inputs = {}
        self.timer_label = None
        self.log_label = None
        self.status_label = None
        self.relation_input = None

        self._build_ui()
        self.load_session()
        Clock.schedule_interval(self._tick, 0.05)

    def _build_ui(self):
        header = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(72), spacing=dp(2))
        header.add_widget(Label(text="VRS Studio", bold=True, font_size="22sp", halign="left"))
        header.add_widget(Label(text="App movil offline", font_size="14sp", halign="left"))
        self.add_widget(header)

        tabs = TabbedPanel(do_default_tab=False, tab_height=dp(42))
        tabs.add_widget(self._session_tab())
        tabs.add_widget(self._timing_tab())
        tabs.add_widget(self._setup_tab())
        tabs.add_widget(self._export_tab())
        self.add_widget(tabs)

        self.status_label = Label(text="Listo", size_hint_y=None, height=dp(30))
        self.add_widget(self.status_label)

    def _session_tab(self):
        tab = TabbedPanelItem(text="Sesion")
        scroll = ScrollView()
        grid = GridLayout(cols=1, spacing=dp(8), padding=(0, dp(8)), size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        self._add_spinner(grid, "Modalidad", "modalidad", MODALITIES)
        self._add_input(grid, "Categoria", "categoria")
        self._add_spinner(grid, "Tipo de sesion", "tipo_sesion", SESSION_TYPES)
        self._add_input(grid, "Campeonato", "campeonato")
        self._add_input(grid, "Equipo", "equipo")
        self._add_input(grid, "Evento", "evento")
        self._add_input(grid, "Circuito", "circuito")
        self._add_input(grid, "Fecha", "fecha")
        self._add_input(grid, "Piloto", "piloto")
        self._add_input(grid, "Dorsal", "dorsal")
        self._add_input(grid, "Mecanico", "mecanico")
        self._add_input(grid, "Chasis", "chasis")
        self._add_input(grid, "Motor", "motor_nombre")
        self._add_spinner(grid, "Clima", "clima", WEATHER_CONDITIONS)
        self._add_multiline_input(grid, "Observaciones", "session_notes", height=dp(120))

        scroll.add_widget(grid)
        tab.add_widget(scroll)
        return tab

    def _timing_tab(self):
        tab = TabbedPanelItem(text="Tiempos")
        scroll = ScrollView()
        container = GridLayout(cols=1, spacing=dp(8), padding=(0, dp(8)), size_hint_y=None)
        container.bind(minimum_height=container.setter("height"))

        timer_box = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(170), spacing=dp(8))
        timer_box.add_widget(Label(text="Cronometro", size_hint_y=None, height=dp(24)))
        self.timer_label = Label(text="00:00.000", font_size="30sp", size_hint_y=None, height=dp(54))
        timer_box.add_widget(self.timer_label)

        row1 = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        row1.add_widget(self._action_button("Iniciar", self.start_timer))
        row1.add_widget(self._action_button("Parar", self.stop_timer))
        row1.add_widget(self._action_button("Reset", self.reset_timer))
        timer_box.add_widget(row1)
        timer_box.add_widget(self._action_button("Registrar vuelta", self.register_lap, height=dp(48)))
        container.add_widget(timer_box)

        self._add_spinner(container, "Estado pista", "estado_pista", TRACK_STATES)
        self._add_spinner(container, "Tipo de evento", "tipo_evento", EVENT_TYPES)
        self._add_spinner(container, "Clima pista", "clima_pista", WEATHER_CONDITIONS)
        self._add_multiline_input(container, "Notas de registro", "timing_notes", height=dp(110))

        actions = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(150))
        actions.add_widget(self._action_button("Anadir evento", self.add_manual_event))
        actions.add_widget(self._action_button("Cambio clima", self.add_weather_change))
        actions.add_widget(self._action_button("Pit Out", self.mark_pit_out))
        actions.add_widget(self._action_button("Pit In", self.mark_pit_in))
        actions.add_widget(self._action_button("Sale SC", lambda *_: self.quick_event("Safety Car Sale", "Safety Car")))
        actions.add_widget(self._action_button("Entra SC", lambda *_: self.quick_event("Safety Car Entra", "Green Flag")))
        container.add_widget(actions)

        container.add_widget(Label(text="Registro de vueltas", size_hint_y=None, height=dp(26)))
        self.log_label = Label(
            text="Todavia no hay registros.",
            size_hint_y=None,
            valign="top",
            halign="left",
            text_size=(dp(320), None),
        )
        self.log_label.bind(texture_size=self._resize_log_label)
        container.add_widget(self.log_label)

        scroll.add_widget(container)
        tab.add_widget(scroll)
        return tab

    def _setup_tab(self):
        tab = TabbedPanelItem(text="Setup")
        scroll = ScrollView()
        grid = GridLayout(cols=1, spacing=dp(8), padding=(0, dp(8)), size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        for label, key in [
            ("Distancia ruedas delante", "distancia_delante"),
            ("Distancia ruedas detras", "distancia_detras"),
            ("Caida izquierda", "caida_izq"),
            ("Caida derecha", "caida_der"),
            ("Caster izquierdo", "caster_izq"),
            ("Caster derecho", "caster_der"),
            ("Reactividad", "reactividad"),
            ("Avance", "avance"),
            ("Tipo de neumatico", "neumatico"),
            ("Tipo de llanta", "llanta"),
            ("Tipo de buje", "buje"),
            ("Ancho trasero", "ancho_trasero"),
            ("Altura de eje", "altura_eje"),
            ("Pinon", "pinon"),
            ("Corona", "corona"),
            ("Relacion", "relacion"),
            ("Tipo de eje", "tipo_eje"),
            ("Presion del. izq.", "presion_del_izq"),
            ("Presion del. der.", "presion_del_der"),
            ("Presion tras. izq.", "presion_tras_izq"),
            ("Presion tras. der.", "presion_tras_der"),
            ("Bujia", "bujia"),
            ("Carburacion", "carburacion"),
            ("Desarrollo", "desarrollo"),
            ("Temperatura motor", "temperatura_motor"),
            ("Aguja altas", "aguja_altas"),
            ("Aguja bajas", "aguja_bajas"),
        ]:
            readonly = key == "relacion"
            widget = self._add_input(grid, label, key, readonly=readonly)
            if key == "relacion":
                self.relation_input = widget

        scroll.add_widget(grid)
        tab.add_widget(scroll)
        return tab

    def _export_tab(self):
        tab = TabbedPanelItem(text="Exportar")
        box = BoxLayout(orientation="vertical", spacing=dp(10), padding=(0, dp(8)))
        box.add_widget(Label(text="Guardado local y exportacion JSON / Excel", size_hint_y=None, height=dp(28)))
        box.add_widget(self._action_button("Guardar sesion", self.save_session))
        box.add_widget(self._action_button("Exportar JSON", self.export_session))
        box.add_widget(self._action_button("Exportar Excel", self.export_excel))
        box.add_widget(self._action_button("Borrar datos", self.reset_all))
        tab.add_widget(box)
        return tab

    def _field_box(self, label_text, widget):
        box = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(84), spacing=dp(4))
        box.add_widget(Label(text=label_text, size_hint_y=None, height=dp(24), halign="left"))
        box.add_widget(widget)
        return box

    def _add_input(self, parent, label, key, readonly=False):
        widget = TextInput(
            text=self.form_values.get(key, ""),
            multiline=False,
            readonly=readonly,
            size_hint_y=None,
            height=dp(44),
        )
        widget.bind(text=lambda instance, value, field=key: self._on_input_change(field, value))
        parent.add_widget(self._field_box(label, widget))
        self.inputs[key] = widget
        return widget

    def _add_multiline_input(self, parent, label, key, height=dp(100)):
        widget = TextInput(
            text=self.form_values.get(key, ""),
            multiline=True,
            size_hint_y=None,
            height=height,
        )
        widget.bind(text=lambda instance, value, field=key: self._on_input_change(field, value))
        parent.add_widget(self._field_box(label, widget))
        self.inputs[key] = widget
        return widget

    def _add_spinner(self, parent, label, key, values):
        widget = Spinner(
            text=self.form_values.get(key, values[0]),
            values=values,
            size_hint_y=None,
            height=dp(44),
        )
        widget.bind(text=lambda instance, value, field=key: self._on_input_change(field, value))
        parent.add_widget(self._field_box(label, widget))
        self.inputs[key] = widget
        return widget

    def _action_button(self, text, callback, height=dp(44)):
        button = Button(text=text, size_hint_y=None, height=height)
        button.bind(on_release=callback)
        return button

    def _on_input_change(self, key, value):
        self.form_values[key] = value
        if key in {"pinon", "corona"}:
            self._update_ratio()

    def _update_ratio(self):
        pinon = self.form_values.get("pinon", "").strip().replace(",", ".")
        corona = self.form_values.get("corona", "").strip().replace(",", ".")
        try:
            pinon_value = float(pinon)
            corona_value = float(corona)
            if pinon_value == 0:
                raise ValueError
            ratio = f"{corona_value / pinon_value:.3f}"
        except ValueError:
            ratio = "-"
        self.form_values["relacion"] = ratio
        if self.relation_input is not None and self.relation_input.text != ratio:
            self.relation_input.text = ratio

    def _tick(self, _dt):
        self.timer_label.text = self._format_duration(self._elapsed_seconds())

    def _elapsed_seconds(self):
        if not self.timing_state["timer_running"] or self.timing_state["start_time"] is None:
            return self.timing_state["elapsed_before_start"]
        return self.timing_state["elapsed_before_start"] + (datetime.now().timestamp() - self.timing_state["start_time"])

    def _format_duration(self, seconds):
        safe = max(float(seconds), 0.0)
        minutes = int(safe // 60)
        remainder = int(safe % 60)
        milliseconds = int((safe - int(safe)) * 1000)
        return f"{minutes:02d}:{remainder:02d}.{milliseconds:03d}"

    def start_timer(self, *_args):
        if self.timing_state["timer_running"]:
            return
        self.timing_state["start_time"] = datetime.now().timestamp()
        self.timing_state["timer_running"] = True
        self._set_status("Cronometro iniciado")

    def stop_timer(self, *_args):
        if not self.timing_state["timer_running"]:
            return
        self.timing_state["elapsed_before_start"] = self._elapsed_seconds()
        self.timing_state["timer_running"] = False
        self.timing_state["start_time"] = None
        self._set_status("Cronometro parado")

    def reset_timer(self, *_args):
        self.timing_state.update(deepcopy(DEFAULT_TIMING))
        self._render_log()
        self._set_status("Cronometro reseteado")

    def register_lap(self, *_args):
        if self.timing_state["pending_lap_status"] == "Pit In":
            self.stop_timer()
        total_elapsed = self._elapsed_seconds()
        lap_elapsed = total_elapsed - self.timing_state["last_lap_elapsed"]
        lap_status = self.timing_state["pending_lap_status"] or self.form_values["estado_pista"]
        lap_event = self.timing_state["pending_lap_status"] or "Vuelta"

        row = {
            "lap": self.timing_state["lap_number"],
            "hora": datetime.now().strftime("%H:%M:%S"),
            "acumulado": self._format_duration(total_elapsed),
            "vuelta": self._format_duration(lap_elapsed),
            "estado": lap_status,
            "clima": self.form_values["clima_pista"],
            "evento": lap_event,
            "notas": self.form_values["timing_notes"].strip(),
        }
        self.timing_state["rows"].append(row)
        self.timing_state["last_lap_elapsed"] = total_elapsed
        self.timing_state["lap_number"] += 1
        self.timing_state["pending_lap_status"] = None
        self._set_input_value("timing_notes", "")
        self._render_log()
        self._set_status(f"Vuelta {row['lap']} registrada")

    def add_manual_event(self, *_args):
        row = {
            "lap": "-",
            "hora": datetime.now().strftime("%H:%M:%S"),
            "acumulado": self._format_duration(self._elapsed_seconds()),
            "vuelta": "-",
            "estado": self.form_values["estado_pista"],
            "clima": self.form_values["clima_pista"],
            "evento": self.form_values["tipo_evento"],
            "notas": self.form_values["timing_notes"].strip(),
        }
        self.timing_state["rows"].append(row)
        self._set_input_value("timing_notes", "")
        self._render_log()
        self._set_status(f"Evento {row['evento']} anadido")

    def add_weather_change(self, *_args):
        note = self.form_values["timing_notes"].strip()
        weather = self.form_values["clima_pista"]
        combined = f"Nuevo clima: {weather}" if not note else f"Nuevo clima: {weather} | {note}"
        row = {
            "lap": "-",
            "hora": datetime.now().strftime("%H:%M:%S"),
            "acumulado": self._format_duration(self._elapsed_seconds()),
            "vuelta": "-",
            "estado": self.form_values["estado_pista"],
            "clima": weather,
            "evento": "Cambio de clima",
            "notas": combined,
        }
        self.form_values["clima"] = weather
        self.timing_state["rows"].append(row)
        self._set_input_value("clima", weather)
        self._set_input_value("timing_notes", "")
        self._render_log()
        self._set_status("Cambio de clima registrado")

    def quick_event(self, event_type, track_state):
        self._set_input_value("tipo_evento", event_type)
        self._set_input_value("estado_pista", track_state)
        self.add_manual_event()

    def mark_pit_out(self, *_args):
        self._set_input_value("tipo_evento", "Pit Out")
        self._set_input_value("estado_pista", "Green Flag")
        self.timing_state["pending_lap_status"] = "Pit Out"
        self.start_timer()
        self._set_status("Pit Out preparado")

    def mark_pit_in(self, *_args):
        self._set_input_value("tipo_evento", "Pit In")
        self._set_input_value("estado_pista", "Pit Lane")
        self.timing_state["pending_lap_status"] = "Pit In"
        self._set_status("Pit In preparado")

    def _render_log(self):
        rows = self.timing_state["rows"]
        if not rows:
            self.log_label.text = "Todavia no hay registros."
            return
        lines = []
        for row in reversed(rows):
            prefix = "Evento" if row["lap"] == "-" else f"Vuelta {row['lap']}"
            lines.append(
                f"{prefix} | {row['hora']} | {row['evento']} | Total {row['acumulado']} | "
                f"Vuelta {row['vuelta']} | Estado {row['estado']} | Clima {row['clima']} | "
                f"Notas: {row['notas'] or '-'}"
            )
        self.log_label.text = "\n\n".join(lines)

    def _resize_log_label(self, instance, size):
        instance.height = max(size[1], dp(80))
        instance.text_size = (instance.width, None)

    def _set_input_value(self, key, value):
        self.form_values[key] = value
        widget = self.inputs.get(key)
        if widget is None:
            return
        if widget.text != value:
            widget.text = value

    def _storage_path(self):
        base = Path(App.get_running_app().user_data_dir)
        base.mkdir(parents=True, exist_ok=True)
        return base / "session_data.json"

    def save_session(self, *_args):
        payload = {
            "form_values": self.form_values,
            "timing_state": self.timing_state,
        }
        self._storage_path().write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
        self._set_status("Sesion guardada en local")

    def load_session(self):
        path = self._storage_path()
        if not path.exists():
            self._render_log()
            self._update_ratio()
            return
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            self.form_values.update(payload.get("form_values", {}))
            self.timing_state.update(payload.get("timing_state", {}))
            self.timing_state["start_time"] = None
            self.timing_state["timer_running"] = False
            for key, value in self.form_values.items():
                if key in self.inputs:
                    self._set_input_value(key, value)
            self._update_ratio()
            self._render_log()
            self._set_status("Sesion cargada")
        except Exception:
            self._render_log()
            self._set_status("No se pudo cargar la sesion")

    def export_session(self, *_args):
        export_dir = Path(App.get_running_app().user_data_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_path = export_dir / f"vrs_mobile_export_{stamp}.json"
        payload = {
            "form_values": self.form_values,
            "timing_state": self.timing_state,
        }
        export_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
        self._set_status(f"Exportado: {export_path.name}")

    def export_excel(self, *_args):
        export_dir = Path(App.get_running_app().user_data_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        category = (self.form_values.get("categoria", "").strip().lower() or "sesion").replace(" ", "_")
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_path = export_dir / f"reporte_{category}_{stamp}.xlsx"

        workbook = Workbook()
        workbook.active.title = "Resumen"
        self._write_summary_sheet(workbook["Resumen"])
        self._write_timing_sheet(workbook.create_sheet("Tiempos"))
        self._write_setup_sheet(workbook.create_sheet("Setup"))
        workbook.save(export_path)
        self._set_status(f"Excel exportado: {export_path.name}")

    def _write_summary_sheet(self, sheet):
        sheet["A1"] = "Resumen de sesion"
        row = 3
        pairs = [
            ("Modalidad", self.form_values["modalidad"]),
            ("Categoria", self.form_values["categoria"]),
            ("Campeonato", self.form_values["campeonato"]),
            ("Equipo", self.form_values["equipo"]),
            ("Tipo de sesion", self.form_values["tipo_sesion"]),
            ("Evento", self.form_values["evento"]),
            ("Circuito", self.form_values["circuito"]),
            ("Fecha", self.form_values["fecha"]),
            ("Piloto", self.form_values["piloto"]),
            ("Dorsal", self.form_values["dorsal"]),
            ("Mecanico", self.form_values["mecanico"]),
            ("Chasis", self.form_values["chasis"]),
            ("Motor", self.form_values["motor_nombre"]),
            ("Clima", self.form_values["clima"]),
        ]
        for label, value in pairs:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            row += 1

        sheet[f"A{row + 1}"] = "Observaciones"
        sheet[f"B{row + 1}"] = self.form_values["session_notes"].strip()
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 60

    def _write_timing_sheet(self, sheet):
        headers = ["Vuelta", "Hora", "Tiempo total", "Tiempo vuelta", "Estado pista", "Clima", "Evento", "Notas"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row_index, row_data in enumerate(self.timing_state["rows"], start=2):
            values = [
                row_data["lap"],
                row_data["hora"],
                row_data["acumulado"],
                row_data["vuelta"],
                row_data["estado"],
                row_data["clima"],
                row_data["evento"],
                row_data["notas"],
            ]
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        widths = [12, 14, 16, 16, 18, 18, 18, 40]
        for col_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + col_index)].width = width

    def _write_setup_sheet(self, sheet):
        rows = [
            ("Set up direccion", ""),
            ("Distancia ruedas delante", self.form_values["distancia_delante"]),
            ("Distancia ruedas detras", self.form_values["distancia_detras"]),
            ("Caida izquierda", self.form_values["caida_izq"]),
            ("Caida derecha", self.form_values["caida_der"]),
            ("Caster izquierdo", self.form_values["caster_izq"]),
            ("Caster derecho", self.form_values["caster_der"]),
            ("Reactividad", self.form_values["reactividad"]),
            ("Avance", self.form_values["avance"]),
            ("", ""),
            ("Set up eje trasero", ""),
            ("Tipo de neumatico", self.form_values["neumatico"]),
            ("Tipo de llanta", self.form_values["llanta"]),
            ("Tipo de buje", self.form_values["buje"]),
            ("Ancho trasero", self.form_values["ancho_trasero"]),
            ("Altura de eje", self.form_values["altura_eje"]),
            ("Pinon", self.form_values["pinon"]),
            ("Corona", self.form_values["corona"]),
            ("Relacion", self.form_values["relacion"]),
            ("Tipo de eje", self.form_values["tipo_eje"]),
            ("Presion delantera izquierda", self.form_values["presion_del_izq"]),
            ("Presion delantera derecha", self.form_values["presion_del_der"]),
            ("Presion trasera izquierda", self.form_values["presion_tras_izq"]),
            ("Presion trasera derecha", self.form_values["presion_tras_der"]),
            ("", ""),
            ("Set up motor", ""),
            ("Bujia", self.form_values["bujia"]),
            ("Carburacion", self.form_values["carburacion"]),
            ("Desarrollo", self.form_values["desarrollo"]),
            ("Temperatura motor", self.form_values["temperatura_motor"]),
            ("Aguja altas", self.form_values["aguja_altas"]),
            ("Aguja bajas", self.form_values["aguja_bajas"]),
        ]
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 40

    def reset_all(self, *_args):
        self.form_values = deepcopy(DEFAULT_FORM)
        self.timing_state = deepcopy(DEFAULT_TIMING)
        for key, widget in self.inputs.items():
            value = self.form_values.get(key, "")
            if widget.text != value:
                widget.text = value
        self._update_ratio()
        self._render_log()
        self.save_session()
        self._set_status("Datos borrados")

    def _set_status(self, text):
        self.status_label.text = text


class VRSMobileApp(App):
    def build(self):
        self.title = "VRS Studio Movil"
        return MobileRoot()


if __name__ == "__main__":
    VRSMobileApp().run()
