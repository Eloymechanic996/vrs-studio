import tkinter as tk
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook


TRACK_STATES = [
    "Green Flag",
    "Yellow Flag",
    "Red Flag",
    "Safety Car",
    "Virtual Safety Car",
    "Pit Lane",
]

EVENT_TYPES = [
    "Incidencia",
    "Pit In",
    "Pit Out",
    "Safety Car Sale",
    "Safety Car Entra",
    "Bandera",
    "Cambio de clima",
]

SESSION_TYPES = [
    "Race",
    "Test",
    "Free Practice",
    "Qualifying",
]

MODALITIES = [
    "Karting",
    "Monoplazas",
    "Rally",
    "Turismos",
    "Endurance",
    "Off-Road",
]

WEATHER_CONDITIONS = [
    "Dry",
    "Cloudy",
    "Light Rain",
    "Rain",
    "Heavy Rain",
    "Windy",
    "Cold",
    "Hot",
]


class EntryRow(ttk.Frame):
    def __init__(self, master, label, variable, width=14):
        super().__init__(master)
        ttk.Label(self, text=label, width=16).pack(side="left", padx=(0, 8))
        ttk.Entry(self, textvariable=variable, width=width).pack(side="left")


class KartingTelemetryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("VRS Studio | Control de Competicion")
        self.root.geometry("1320x860")
        self.root.minsize(1120, 760)

        self.timer_running = False
        self.start_time = None
        self.elapsed_before_start = 0.0
        self.last_lap_elapsed = 0.0
        self.lap_number = 1
        self.timer_job = None
        self.pending_lap_status = None

        self.vars = self._build_variables()
        self.main_container = None
        self.summary_label = None
        self.section_frames = {}
        self.nav_buttons = {}
        self.current_section = None

        self._build_start_screen()
        self._update_ratio()
        self._sync_session_name()

    def _build_variables(self):
        today = datetime.now().strftime("%Y-%m-%d")
        return {
            "modalidad": tk.StringVar(value="Karting"),
            "categoria": tk.StringVar(),
            "tipo_sesion": tk.StringVar(value="Test"),
            "campeonato": tk.StringVar(),
            "equipo": tk.StringVar(),
            "evento": tk.StringVar(),
            "circuito": tk.StringVar(),
            "fecha": tk.StringVar(value=today),
            "piloto": tk.StringVar(),
            "dorsal": tk.StringVar(),
            "mecanico": tk.StringVar(value="Jose"),
            "chasis": tk.StringVar(),
            "motor_nombre": tk.StringVar(),
            "clima": tk.StringVar(),
            "clima_pista": tk.StringVar(value="Dry"),
            "plantilla_excel": tk.StringVar(),
            "estado_pista": tk.StringVar(value=TRACK_STATES[0]),
            "tipo_evento": tk.StringVar(value="Incidencia"),
            "nota_evento": tk.StringVar(),
            "distancia_delante": tk.StringVar(),
            "distancia_detras": tk.StringVar(),
            "caida_izq": tk.StringVar(),
            "caida_der": tk.StringVar(),
            "caster_izq": tk.StringVar(),
            "caster_der": tk.StringVar(),
            "reactividad": tk.StringVar(),
            "avance": tk.StringVar(),
            "neumatico": tk.StringVar(),
            "llanta": tk.StringVar(),
            "buje": tk.StringVar(),
            "ancho_trasero": tk.StringVar(),
            "altura_eje": tk.StringVar(),
            "presion_del_izq": tk.StringVar(),
            "presion_del_der": tk.StringVar(),
            "presion_tras_izq": tk.StringVar(),
            "presion_tras_der": tk.StringVar(),
            "pinon": tk.StringVar(),
            "corona": tk.StringVar(),
            "relacion": tk.StringVar(value="-"),
            "tipo_eje": tk.StringVar(),
            "bujia": tk.StringVar(),
            "carburacion": tk.StringVar(),
            "desarrollo": tk.StringVar(),
            "temperatura_motor": tk.StringVar(),
            "aguja_altas": tk.StringVar(),
            "aguja_bajas": tk.StringVar(),
        }

    def _build_start_screen(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(".", background="#f3f1eb", foreground="#1f2933")
        style.configure("TFrame", background="#f3f1eb")
        style.configure("TLabelframe", background="#fbfaf7", borderwidth=1, relief="solid")
        style.configure("TLabelframe.Label", background="#fbfaf7", foreground="#24413b")
        style.configure("TLabel", background="#f3f1eb", foreground="#1f2933", font=("Segoe UI", 10))
        style.configure("Title.TLabel", font=("Bahnschrift", 20, "bold"), foreground="#16302b")
        style.configure("StartTitle.TLabel", font=("Bahnschrift", 28, "bold"))
        style.configure("Section.TLabelframe", background="#fbfaf7", bordercolor="#d7d0c4")
        style.configure("Section.TLabelframe.Label", font=("Bahnschrift", 11, "bold"), foreground="#24413b")
        style.configure("Timer.TLabel", font=("Consolas", 30, "bold"))
        style.configure("TButton", font=("Segoe UI", 10), padding=(10, 6))
        style.configure("TEntry", padding=4)
        style.configure("TCombobox", padding=3)
        style.configure("Sidebar.TFrame", background="#16302b")
        style.configure("SidebarTitle.TLabel", background="#16302b", foreground="#f4efe4", font=("Bahnschrift", 18, "bold"))
        style.configure("SidebarInfo.TLabel", background="#16302b", foreground="#dbe6dd", font=("Segoe UI", 10))
        style.configure("Nav.TButton", font=("Bahnschrift", 11), padding=(12, 10))
        style.map(
            "Nav.TButton",
            background=[("active", "#d4b16a"), ("!disabled", "#f1e2bd")],
            foreground=[("active", "#16302b"), ("!disabled", "#16302b")],
        )
        style.configure("ActiveNav.TButton", font=("Bahnschrift", 11, "bold"), padding=(12, 10), background="#c9953d", foreground="#ffffff")

        self.start_container = ttk.Frame(self.root, padding=28)
        self.start_container.pack(fill="both", expand=True)

        shell = ttk.Frame(self.start_container)
        shell.place(relx=0.5, rely=0.5, anchor="center")

        ttk.Label(shell, text="Menu inicial de competicion", style="StartTitle.TLabel").pack(anchor="center")
        ttk.Label(
            shell,
            text="Completa los datos del coche y de la sesion antes de entrar al panel de trabajo.",
        ).pack(anchor="center", pady=(6, 18))

        card = ttk.LabelFrame(shell, text="Datos de salida", padding=22, style="Section.TLabelframe")
        card.pack(fill="x")

        grid = ttk.Frame(card)
        grid.pack(fill="x")
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(3, weight=1)

        starters = [
            ("Numero de chasis", "chasis", "entry"),
            ("Dorsal", "dorsal", "entry"),
            ("Piloto", "piloto", "entry"),
            ("Mecanico a cargo", "mecanico", "entry"),
            ("Equipo", "equipo", "entry"),
            ("Campeonato", "campeonato", "entry"),
            ("Tipo de sesion", "tipo_sesion", "combo"),
            ("Categoria", "categoria", "entry"),
            ("Modalidad", "modalidad", "combo"),
            ("Fecha", "fecha", "entry"),
        ]

        for index, (label, key, kind) in enumerate(starters):
            row = index // 2
            column = (index % 2) * 2
            ttk.Label(grid, text=label).grid(row=row, column=column, sticky="w", padx=(0, 10), pady=8)
            if kind == "combo":
                values = SESSION_TYPES if key == "tipo_sesion" else MODALITIES
                widget = ttk.Combobox(grid, textvariable=self.vars[key], values=values, state="readonly", width=22)
            else:
                widget = ttk.Entry(grid, textvariable=self.vars[key], width=24)
            widget.grid(row=row, column=column + 1, sticky="ew", pady=8)

        footer = ttk.Frame(shell)
        footer.pack(fill="x", pady=(18, 0))
        ttk.Label(
            footer,
            text="Estos datos se cargaran automaticamente en la ficha de sesion y en futuras exportaciones.",
        ).pack(side="left")
        ttk.Button(footer, text="Entrar en la app", command=self.launch_main_app).pack(side="right")

    def launch_main_app(self):
        self.start_container.destroy()
        self._build_main_layout()
        self._update_timer_display()

    def _build_main_layout(self):
        self.main_container = ttk.Frame(self.root, padding=16)
        self.main_container.pack(fill="both", expand=True)

        shell = ttk.Frame(self.main_container)
        shell.pack(fill="both", expand=True)

        sidebar = ttk.Frame(shell, style="Sidebar.TFrame", width=250, padding=18)
        sidebar.pack(side="left", fill="y", padx=(0, 14))
        sidebar.pack_propagate(False)

        ttk.Label(sidebar, text="VRS Studio", style="SidebarTitle.TLabel").pack(anchor="w", pady=(0, 6))
        ttk.Label(sidebar, text="Control de sesion y setup", style="SidebarInfo.TLabel").pack(anchor="w", pady=(0, 18))

        self.summary_label = ttk.Label(sidebar, text="", style="SidebarInfo.TLabel", wraplength=190, justify="left")
        self.summary_label.pack(anchor="w", fill="x", pady=(0, 18))
        self._refresh_summary_label()

        nav_items = [
            ("Sesion", "session"),
            ("Tiempos", "timing"),
            ("Direccion", "front"),
            ("Eje trasero", "rear"),
            ("Motor", "engine"),
        ]
        for label, key in nav_items:
            button = ttk.Button(sidebar, text=label, style="Nav.TButton", command=lambda section=key: self.show_section(section))
            button.pack(fill="x", pady=4)
            self.nav_buttons[key] = button

        ttk.Frame(sidebar, style="Sidebar.TFrame").pack(fill="y", expand=True)
        ttk.Button(sidebar, text="Editar datos iniciales", command=self.edit_start_data).pack(fill="x", pady=(0, 8))
        ttk.Button(sidebar, text="Exportar a Excel", command=self.export_to_excel).pack(fill="x")

        content_shell = ttk.Frame(shell)
        content_shell.pack(side="left", fill="both", expand=True)

        header_card = ttk.LabelFrame(content_shell, text="Panel de trabajo", padding=14, style="Section.TLabelframe")
        header_card.pack(fill="x", pady=(0, 12))
        ttk.Label(header_card, text="Registro de setup, pista y competicion", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            header_card,
            text="Navega desde el lateral para cambiar de apartado y mantener la informacion ordenada durante la sesion.",
        ).pack(anchor="w", pady=(4, 0))

        content_area = ttk.Frame(content_shell)
        content_area.pack(fill="both", expand=True)

        self.session_tab = ttk.Frame(content_area, padding=14)
        self.timing_tab = ttk.Frame(content_area, padding=14)
        self.front_tab = ttk.Frame(content_area, padding=14)
        self.rear_tab = ttk.Frame(content_area, padding=14)
        self.engine_tab = ttk.Frame(content_area, padding=14)

        self.section_frames = {
            "session": self.session_tab,
            "timing": self.timing_tab,
            "front": self.front_tab,
            "rear": self.rear_tab,
            "engine": self.engine_tab,
        }

        for frame in self.section_frames.values():
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._build_session_tab()
        self._build_timing_tab()
        self._build_front_tab()
        self._build_rear_tab()
        self._build_engine_tab()
        self.show_section("session")

    def show_section(self, section_key):
        if section_key not in self.section_frames:
            return
        self.section_frames[section_key].tkraise()
        self.current_section = section_key
        for key, button in self.nav_buttons.items():
            button.configure(style="ActiveNav.TButton" if key == section_key else "Nav.TButton")

    def edit_start_data(self):
        self.root.destroy()
        main()

    def _build_session_tab(self):
        info = ttk.LabelFrame(self.session_tab, text="Datos generales", padding=16, style="Section.TLabelframe")
        info.pack(fill="x", pady=(0, 12))

        grid = ttk.Frame(info)
        grid.pack(fill="x")
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(3, weight=1)

        fields = [
            ("Modalidad", "modalidad"),
            ("Categoria", "categoria"),
            ("Tipo de sesion", "tipo_sesion"),
            ("Equipo", "equipo"),
            ("Campeonato", "campeonato"),
            ("Evento", "evento"),
            ("Circuito", "circuito"),
            ("Fecha", "fecha"),
            ("Piloto", "piloto"),
            ("Dorsal", "dorsal"),
            ("Mecanico", "mecanico"),
            ("Chasis", "chasis"),
            ("Motor", "motor_nombre"),
            ("Clima", "clima"),
        ]

        for index, (label, key) in enumerate(fields):
            row = index // 2
            column = (index % 2) * 2
            ttk.Label(grid, text=label).grid(row=row, column=column, sticky="w", padx=(0, 10), pady=6)
            widget = ttk.Entry(grid, textvariable=self.vars[key], width=24)
            if key in {"modalidad", "tipo_sesion"}:
                widget.destroy()
                values = MODALITIES if key == "modalidad" else SESSION_TYPES
                widget = ttk.Combobox(grid, textvariable=self.vars[key], values=values, state="readonly", width=22)
            widget.grid(row=row, column=column + 1, sticky="ew", pady=6)

        template_frame = ttk.LabelFrame(self.session_tab, text="Plantilla Excel", padding=16, style="Section.TLabelframe")
        template_frame.pack(fill="x", pady=(0, 12))
        template_frame.columnconfigure(1, weight=1)
        ttk.Label(template_frame, text="Archivo base .xlsx").grid(row=0, column=0, sticky="w", padx=(0, 10))
        ttk.Entry(template_frame, textvariable=self.vars["plantilla_excel"]).grid(row=0, column=1, sticky="ew")
        ttk.Button(template_frame, text="Buscar", command=self.choose_template).grid(row=0, column=2, padx=(10, 0))
        ttk.Label(
            template_frame,
            text="Si eliges una plantilla, la app rellenara sus hojas Resumen, Tiempos y Setup. "
            "Si no existe, se creara un Excel nuevo.",
            wraplength=900,
            justify="left",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 0))

        notes_frame = ttk.LabelFrame(self.session_tab, text="Observaciones generales", padding=16, style="Section.TLabelframe")
        notes_frame.pack(fill="both", expand=True)
        self.session_notes = tk.Text(notes_frame, height=24, wrap="word", bd=1, relief="solid")
        self.session_notes.pack(fill="both", expand=True)

    def _build_timing_tab(self):
        top = ttk.Frame(self.timing_tab)
        top.pack(fill="x", pady=(0, 12))

        timer_box = ttk.LabelFrame(top, text="Cronometro", padding=16, style="Section.TLabelframe")
        timer_box.pack(side="left", fill="y")
        self.timer_label = ttk.Label(timer_box, text="00:00.000", style="Timer.TLabel")
        self.timer_label.pack(anchor="w")

        timer_actions = ttk.Frame(timer_box)
        timer_actions.pack(anchor="w", pady=(12, 0))
        ttk.Button(timer_actions, text="Iniciar", command=self.start_timer).pack(side="left", padx=(0, 6))
        ttk.Button(timer_actions, text="Parar", command=self.stop_timer).pack(side="left", padx=6)
        ttk.Button(timer_actions, text="Reset", command=self.reset_timer).pack(side="left", padx=6)
        ttk.Button(timer_actions, text="Registrar vuelta", command=self.register_lap).pack(side="left", padx=6)

        event_box = ttk.LabelFrame(top, text="Evento en pista", padding=16, style="Section.TLabelframe")
        event_box.pack(side="left", fill="both", expand=True, padx=(12, 0))
        event_box.columnconfigure(1, weight=1)
        event_box.columnconfigure(3, weight=1)

        ttk.Label(event_box, text="Estado pista").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Combobox(
            event_box,
            textvariable=self.vars["estado_pista"],
            values=TRACK_STATES,
            state="readonly",
        ).grid(row=0, column=1, sticky="ew", pady=6)

        ttk.Label(event_box, text="Tipo de evento").grid(row=0, column=2, sticky="w", padx=(18, 10), pady=6)
        ttk.Combobox(
            event_box,
            textvariable=self.vars["tipo_evento"],
            values=EVENT_TYPES,
            state="readonly",
        ).grid(row=0, column=3, sticky="ew", pady=6)

        ttk.Label(event_box, text="Clima").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Combobox(
            event_box,
            textvariable=self.vars["clima_pista"],
            values=WEATHER_CONDITIONS,
            state="readonly",
        ).grid(row=1, column=1, sticky="ew", pady=6)

        actions = ttk.Frame(event_box)
        actions.grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 0))
        ttk.Button(actions, text="Anadir evento", command=self.add_manual_event).pack(side="left", padx=(0, 6))
        ttk.Button(actions, text="Pit In", command=self.mark_pit_in).pack(side="left", padx=6)
        ttk.Button(actions, text="Pit Out", command=self.mark_pit_out).pack(side="left", padx=6)
        ttk.Button(actions, text="Sale SC", command=lambda: self.quick_event("Safety Car Sale", "Safety Car")).pack(side="left", padx=6)
        ttk.Button(actions, text="Entra SC", command=lambda: self.quick_event("Safety Car Entra", "Green Flag")).pack(side="left", padx=6)
        ttk.Button(actions, text="Cambio clima", command=self.add_weather_change).pack(side="left", padx=6)

        notes_box = ttk.LabelFrame(self.timing_tab, text="Notas del registro", padding=12, style="Section.TLabelframe")
        notes_box.pack(fill="x", pady=(0, 12))
        self.timing_notes = tk.Text(notes_box, height=3, wrap="word", bd=1, relief="solid")
        self.timing_notes.pack(fill="x", expand=True)

        table_box = ttk.LabelFrame(self.timing_tab, text="Registro de vueltas y pista", padding=12, style="Section.TLabelframe")
        table_box.pack(fill="both", expand=True)

        columns = ("lap", "hora", "acumulado", "vuelta", "estado", "clima", "evento", "notas")
        self.timing_table = ttk.Treeview(table_box, columns=columns, show="headings", height=18)
        headings = {
            "lap": "Vuelta",
            "hora": "Hora",
            "acumulado": "Tiempo total",
            "vuelta": "Tiempo vuelta",
            "estado": "Estado pista",
            "clima": "Clima",
            "evento": "Evento",
            "notas": "Notas",
        }
        widths = {"lap": 70, "hora": 110, "acumulado": 120, "vuelta": 110, "estado": 120, "clima": 120, "evento": 130, "notas": 250}
        for key in columns:
            self.timing_table.heading(key, text=headings[key])
            self.timing_table.column(key, width=widths[key], anchor="center" if key != "notas" else "w")

        scrollbar = ttk.Scrollbar(table_box, orient="vertical", command=self.timing_table.yview)
        self.timing_table.configure(yscrollcommand=scrollbar.set)
        self.timing_table.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _build_front_tab(self):
        frame = ttk.LabelFrame(self.front_tab, text="Set up de direccion", padding=16, style="Section.TLabelframe")
        frame.pack(anchor="nw", fill="none")

        left = ttk.Frame(frame)
        left.pack(anchor="nw")
        EntryRow(left, "Distancia ruedas delante", self.vars["distancia_delante"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Distancia ruedas detras", self.vars["distancia_detras"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Caida izquierda", self.vars["caida_izq"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Caida derecha", self.vars["caida_der"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Caster izquierdo", self.vars["caster_izq"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Caster derecho", self.vars["caster_der"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Reactividad", self.vars["reactividad"], width=12).pack(anchor="w", pady=6)
        EntryRow(left, "Avance", self.vars["avance"], width=12).pack(anchor="w", pady=6)

    def _build_rear_tab(self):
        frame = ttk.LabelFrame(self.rear_tab, text="Set up eje trasero", padding=16, style="Section.TLabelframe")
        frame.pack(anchor="nw", fill="none")

        EntryRow(frame, "Tipo de neumatico", self.vars["neumatico"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Tipo de llanta", self.vars["llanta"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Tipo de buje", self.vars["buje"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Ancho trasero", self.vars["ancho_trasero"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Altura de eje", self.vars["altura_eje"], width=12).pack(anchor="w", pady=6)

        relation = ttk.Frame(frame)
        relation.pack(anchor="w", pady=6)
        ttk.Label(relation, text="Dientes pinon", width=16).pack(side="left", padx=(0, 8))
        pinon_entry = ttk.Entry(relation, textvariable=self.vars["pinon"], width=10)
        pinon_entry.pack(side="left")
        ttk.Label(relation, text="Dientes corona", width=14).pack(side="left", padx=(16, 8))
        corona_entry = ttk.Entry(relation, textvariable=self.vars["corona"], width=10)
        corona_entry.pack(side="left")
        ttk.Label(relation, text="Relacion", width=10).pack(side="left", padx=(16, 8))
        ttk.Label(relation, textvariable=self.vars["relacion"]).pack(side="left")

        self.vars["pinon"].trace_add("write", lambda *_: self._update_ratio())
        self.vars["corona"].trace_add("write", lambda *_: self._update_ratio())

        EntryRow(frame, "Tipo de eje", self.vars["tipo_eje"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Presion del. izq.", self.vars["presion_del_izq"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Presion del. der.", self.vars["presion_del_der"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Presion tras. izq.", self.vars["presion_tras_izq"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Presion tras. der.", self.vars["presion_tras_der"], width=12).pack(anchor="w", pady=6)

    def _build_engine_tab(self):
        frame = ttk.LabelFrame(self.engine_tab, text="Set up motor", padding=16, style="Section.TLabelframe")
        frame.pack(anchor="nw", fill="none")

        EntryRow(frame, "Bujia", self.vars["bujia"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Carburacion", self.vars["carburacion"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Desarrollo", self.vars["desarrollo"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Temp. motor", self.vars["temperatura_motor"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Aguja altas", self.vars["aguja_altas"], width=12).pack(anchor="w", pady=6)
        EntryRow(frame, "Aguja bajas", self.vars["aguja_bajas"], width=12).pack(anchor="w", pady=6)

    def choose_template(self):
        path = filedialog.askopenfilename(
            title="Seleccionar plantilla Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if path:
            self.vars["plantilla_excel"].set(path)

    def _elapsed_seconds(self):
        if not self.timer_running or self.start_time is None:
            return self.elapsed_before_start
        return self.elapsed_before_start + (datetime.now().timestamp() - self.start_time)

    def _format_duration(self, seconds):
        delta = timedelta(seconds=max(seconds, 0))
        total_minutes = delta.seconds // 60 + delta.days * 24 * 60
        remainder_seconds = delta.seconds % 60
        milliseconds = int(delta.microseconds / 1000)
        return f"{total_minutes:02d}:{remainder_seconds:02d}.{milliseconds:03d}"

    def _update_timer_display(self):
        self.timer_label.config(text=self._format_duration(self._elapsed_seconds()))
        self.timer_job = self.root.after(50, self._update_timer_display)

    def start_timer(self):
        if self.timer_running:
            return
        self.start_time = datetime.now().timestamp()
        self.timer_running = True

    def stop_timer(self):
        if not self.timer_running:
            return
        self.elapsed_before_start = self._elapsed_seconds()
        self.timer_running = False
        self.start_time = None

    def reset_timer(self):
        self.timer_running = False
        self.start_time = None
        self.elapsed_before_start = 0.0
        self.last_lap_elapsed = 0.0
        self.lap_number = 1
        self.pending_lap_status = None

    def register_lap(self):
        if self.pending_lap_status == "Pit In":
            self.stop_timer()
        total_elapsed = self._elapsed_seconds()
        lap_elapsed = total_elapsed - self.last_lap_elapsed
        lap_status = self.pending_lap_status or self.vars["estado_pista"].get()
        lap_event = self.pending_lap_status or "Vuelta"
        self._insert_timing_row(
            lap=self.lap_number,
            total_time=self._format_duration(total_elapsed),
            lap_time=self._format_duration(lap_elapsed),
            state=lap_status,
            weather=self.vars["clima_pista"].get(),
            event_type=lap_event,
            notes=self._get_timing_notes(),
        )
        self.last_lap_elapsed = total_elapsed
        self.lap_number += 1
        self.pending_lap_status = None
        self._clear_timing_notes()

    def add_manual_event(self):
        self._insert_timing_row(
            lap="-",
            total_time=self._format_duration(self._elapsed_seconds()),
            lap_time="-",
            state=self.vars["estado_pista"].get(),
            weather=self.vars["clima_pista"].get(),
            event_type=self.vars["tipo_evento"].get(),
            notes=self._get_timing_notes(),
        )
        self._clear_timing_notes()

    def add_weather_change(self):
        weather = self.vars["clima_pista"].get()
        note = self._get_timing_notes()
        combined_note = f"Nuevo clima: {weather}" if not note else f"Nuevo clima: {weather} | {note}"
        self._insert_timing_row(
            lap="-",
            total_time=self._format_duration(self._elapsed_seconds()),
            lap_time="-",
            state=self.vars["estado_pista"].get(),
            weather=weather,
            event_type="Cambio de clima",
            notes=combined_note,
        )
        self.vars["clima"].set(weather)
        self._clear_timing_notes()

    def quick_event(self, event_type, track_state):
        self.vars["tipo_evento"].set(event_type)
        self.vars["estado_pista"].set(track_state)
        self.add_manual_event()

    def mark_pit_out(self):
        self.vars["tipo_evento"].set("Pit Out")
        self.vars["estado_pista"].set("Green Flag")
        self.pending_lap_status = "Pit Out"
        self.start_timer()

    def mark_pit_in(self):
        self.vars["tipo_evento"].set("Pit In")
        self.vars["estado_pista"].set("Pit Lane")
        self.pending_lap_status = "Pit In"

    def _insert_timing_row(self, lap, total_time, lap_time, state, weather, event_type, notes):
        now_text = datetime.now().strftime("%H:%M:%S")
        self.timing_table.insert("", "end", values=(lap, now_text, total_time, lap_time, state, weather, event_type, notes))

    def _get_timing_notes(self):
        if not hasattr(self, "timing_notes"):
            return ""
        return self.timing_notes.get("1.0", "end").strip()

    def _clear_timing_notes(self):
        if hasattr(self, "timing_notes"):
            self.timing_notes.delete("1.0", "end")

    def _update_ratio(self):
        pinon = self.vars["pinon"].get().strip()
        corona = self.vars["corona"].get().strip()
        try:
            pinon_value = float(pinon.replace(",", "."))
            corona_value = float(corona.replace(",", "."))
            if pinon_value == 0:
                raise ValueError
            ratio = corona_value / pinon_value
            self.vars["relacion"].set(f"{ratio:.3f}")
        except ValueError:
            self.vars["relacion"].set("-")

    def _collect_timing_rows(self):
        rows = []
        for item in self.timing_table.get_children():
            rows.append(self.timing_table.item(item, "values"))
        return rows

    def _sheet(self, workbook, title):
        if title in workbook.sheetnames:
            sheet = workbook[title]
            sheet.delete_rows(1, sheet.max_row)
            return sheet
        return workbook.create_sheet(title)

    def export_to_excel(self):
        save_path = filedialog.asksaveasfilename(
            title="Guardar informe",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{self.vars['categoria'].get().lower()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )
        if not save_path:
            return

        template_value = self.vars["plantilla_excel"].get().strip()
        try:
            if template_value and Path(template_value).exists():
                workbook = load_workbook(template_value)
            else:
                workbook = Workbook()
                workbook.active.title = "Resumen"

            summary_sheet = self._sheet(workbook, "Resumen")
            timing_sheet = self._sheet(workbook, "Tiempos")
            setup_sheet = self._sheet(workbook, "Setup")

            self._write_summary_sheet(summary_sheet)
            self._write_timing_sheet(timing_sheet)
            self._write_setup_sheet(setup_sheet)

            workbook.save(save_path)
            messagebox.showinfo("Exportacion completada", f"Excel generado correctamente:\n{save_path}")
        except Exception as exc:
            messagebox.showerror("Error al exportar", f"No se pudo generar el Excel.\n\nDetalle: {exc}")

    def _write_summary_sheet(self, sheet):
        sheet["A1"] = "Resumen de sesion"
        row = 3
        pairs = [
            ("Modalidad", self.vars["modalidad"].get()),
            ("Categoria", self.vars["categoria"].get()),
            ("Campeonato", self.vars["campeonato"].get()),
            ("Equipo", self.vars["equipo"].get()),
            ("Tipo de sesion", self.vars["tipo_sesion"].get()),
            ("Evento", self.vars["evento"].get()),
            ("Circuito", self.vars["circuito"].get()),
            ("Fecha", self.vars["fecha"].get()),
            ("Piloto", self.vars["piloto"].get()),
            ("Dorsal", self.vars["dorsal"].get()),
            ("Mecanico", self.vars["mecanico"].get()),
            ("Chasis", self.vars["chasis"].get()),
            ("Motor", self.vars["motor_nombre"].get()),
            ("Clima", self.vars["clima"].get()),
        ]
        for label, value in pairs:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            row += 1

        sheet[f"A{row + 1}"] = "Observaciones"
        sheet[f"B{row + 1}"] = self.session_notes.get("1.0", "end").strip()
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 60

    def _write_timing_sheet(self, sheet):
        headers = ["Vuelta", "Hora", "Tiempo total", "Tiempo vuelta", "Estado pista", "Clima", "Evento", "Notas"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row_index, row_data in enumerate(self._collect_timing_rows(), start=2):
            for col_index, value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        widths = [12, 14, 16, 16, 18, 18, 18, 40]
        for col_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + col_index)].width = width

    def _write_setup_sheet(self, sheet):
        rows = [
            ("Set up direccion", ""),
            ("Distancia ruedas delante", self.vars["distancia_delante"].get()),
            ("Distancia ruedas detras", self.vars["distancia_detras"].get()),
            ("Caida izquierda", self.vars["caida_izq"].get()),
            ("Caida derecha", self.vars["caida_der"].get()),
            ("Caster izquierdo", self.vars["caster_izq"].get()),
            ("Caster derecho", self.vars["caster_der"].get()),
            ("Reactividad", self.vars["reactividad"].get()),
            ("Avance", self.vars["avance"].get()),
            ("", ""),
            ("Set up eje trasero", ""),
            ("Tipo de neumatico", self.vars["neumatico"].get()),
            ("Tipo de llanta", self.vars["llanta"].get()),
            ("Tipo de buje", self.vars["buje"].get()),
            ("Ancho trasero", self.vars["ancho_trasero"].get()),
            ("Altura de eje", self.vars["altura_eje"].get()),
            ("Pinon", self.vars["pinon"].get()),
            ("Corona", self.vars["corona"].get()),
            ("Relacion", self.vars["relacion"].get()),
            ("Tipo de eje", self.vars["tipo_eje"].get()),
            ("Presion delantera izquierda", self.vars["presion_del_izq"].get()),
            ("Presion delantera derecha", self.vars["presion_del_der"].get()),
            ("Presion trasera izquierda", self.vars["presion_tras_izq"].get()),
            ("Presion trasera derecha", self.vars["presion_tras_der"].get()),
            ("", ""),
            ("Set up motor", ""),
            ("Bujia", self.vars["bujia"].get()),
            ("Carburacion", self.vars["carburacion"].get()),
            ("Desarrollo", self.vars["desarrollo"].get()),
            ("Temperatura motor", self.vars["temperatura_motor"].get()),
            ("Aguja altas", self.vars["aguja_altas"].get()),
            ("Aguja bajas", self.vars["aguja_bajas"].get()),
        ]
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 40

    def _sync_session_name(self):
        watched_keys = ["modalidad", "categoria", "tipo_sesion", "campeonato", "equipo", "piloto", "dorsal", "chasis"]
        for key in watched_keys:
            self.vars[key].trace_add("write", lambda *_: self._refresh_summary_label())

    def _refresh_summary_label(self):
        if self.summary_label is None:
            return
        text = (
            f"{self.vars['modalidad'].get() or '-'} | "
            f"{self.vars['categoria'].get() or '-'} | "
            f"{self.vars['campeonato'].get() or '-'} | "
            f"{self.vars['equipo'].get() or '-'} | "
            f"{self.vars['tipo_sesion'].get() or '-'} | "
            f"Piloto {self.vars['piloto'].get() or '-'} | "
            f"Dorsal {self.vars['dorsal'].get() or '-'} | "
            f"Chasis {self.vars['chasis'].get() or '-'}"
        )
        self.summary_label.config(text=text)


def main():
    root = tk.Tk()
    KartingTelemetryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
